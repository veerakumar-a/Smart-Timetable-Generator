
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, colorchooser, filedialog
import sqlite3
import json
import random
import datetime
import threading
import os
from collections import defaultdict
from copy import deepcopy

# Optional libraries with graceful degradation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import HexColor
    HAS_REPORTLAB = True
except ImportError:
  
    HAS_REPORTLAB = False

# CONFIGURATION & CONSTANTS

DB_FILE = "timetable_generator.db"
EXPORTS_DIR = "exports"

# Modern color palette for subjects
COLOR_PALETTE = [
    "#FF6B6B", "#4ECDC4", "#45B7D1", "#FFA07A", "#98D8C8",
    "#F7DC6F", "#BB8FCE", "#85C1E2", "#F8B195", "#C06C84",
    "#6C5B7B", "#355C7D", "#F67280", "#C8E6C9", "#FFE082",
    "#BCAAA4", "#90CAF9", "#CE93D8", "#80CBC4", "#FFAB91"
]

# UI Color scheme
UI_COLORS = {
    'primary': "#FF0000",
    'secondary': '#34495E',
    'accent': '#3498DB',
    'success': '#27AE60',
    'warning': '#F39C12',
    'danger': '#E74C3C',
    'light': '#ECF0F1',
    'dark': '#2C3E50',
    'bg_main': '#F8F9FA',
    'bg_secondary': '#FFFFFF',
    'text_primary': '#2C3E50',
    'text_secondary': '#7F8C8D',
    'border': '#BDC3C7'
}

os.makedirs(EXPORTS_DIR, exist_ok=True)

# UTILITY FUNCTIONS

def credit_to_periods(credit: int) -> int:
    """
    Convert credit hours to periods with specified mapping:
    1 credit -> 3 periods
    3 credits -> 5 periods
    4 credits -> 7 periods
    Others -> credit * 2 periods
    """
    mapping = {1: 3, 3: 5, 4: 7}
    return mapping.get(credit, max(1, credit * 2))

def validate_integer(value, min_val=0, max_val=None):
    """Validate and return integer value within bounds"""
    try:
        val = int(value)
        if val < min_val:
            return min_val
        if max_val and val > max_val:
            return max_val
        return val
    except (ValueError, TypeError):
        return min_val

def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple"""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

# DATABASE MANAGER

class DatabaseManager:
    """Handles all database operations with error handling"""
    
    def __init__(self, db_file=DB_FILE):
        self.db_file = db_file
        self.initialize_db()
    
    def initialize_db(self):
        """Create database tables if they don't exist"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS classes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS subjects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    teacher TEXT,
                    credit INTEGER DEFAULT 0,
                    is_lab INTEGER DEFAULT 0,
                    lab_continuous INTEGER DEFAULT 1,
                    rooms TEXT,
                    color TEXT,
                    FOREIGN KEY(class_id) REFERENCES classes(id) ON DELETE CASCADE
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS timetables (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_id INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    data TEXT NOT NULL,
                    FOREIGN KEY(class_id) REFERENCES classes(id) ON DELETE CASCADE
                )
            ''')
            
            conn.commit()
            conn.close()
        except sqlite3.Error as e:
            print(f"Database initialization error: {e}")
    
    def load_all_data(self):
        """Load all classes and subjects from database"""
        data = {'classes': {}, 'colors': {}}
        
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, name FROM classes")
            classes = cursor.fetchall()
            
            for class_id, class_name in classes:
                data['classes'][class_name] = {
                    'subjects': {},
                    'morning': set(),
                    'afternoon': set()
                }
                
                cursor.execute("""
                    SELECT name, teacher, credit, is_lab, lab_continuous, rooms, color
                    FROM subjects WHERE class_id = ?
                """, (class_id,))
                
                subjects = cursor.fetchall()
                for name, teacher, credit, is_lab, lab_cont, rooms, color in subjects:
                    rooms_list = json.loads(rooms) if rooms else []
                    data['classes'][class_name]['subjects'][name] = {
                        'teacher': teacher or '',
                        'credit': credit,
                        'lab': bool(is_lab),
                        'lab_continuous': lab_cont,
                        'rooms': rooms_list
                    }
                    if color:
                        data['colors'][name] = color
            
            conn.close()
        except sqlite3.Error as e:
            print(f"Error loading data: {e}")
        
        return data
    
    def save_all_data(self, data):
        """Save all classes and subjects to database"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # Clear existing data
            cursor.execute("DELETE FROM subjects")
            cursor.execute("DELETE FROM classes")
            
            # Insert classes and subjects
            for class_name, class_info in data['classes'].items():
                cursor.execute("INSERT INTO classes (name) VALUES (?)", (class_name,))
                class_id = cursor.lastrowid
                
                for subject_name, subject_data in class_info['subjects'].items():
                    rooms_json = json.dumps(subject_data.get('rooms', []))
                    color = data['colors'].get(subject_name, '')
                    
                    cursor.execute("""
                        INSERT INTO subjects 
                        (class_id, name, teacher, credit, is_lab, lab_continuous, rooms, color)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        class_id,
                        subject_name,
                        subject_data.get('teacher', ''),
                        subject_data.get('credit', 0),
                        1 if subject_data.get('lab') else 0,
                        subject_data.get('lab_continuous', 1),
                        rooms_json,
                        color
                    ))
            
            conn.commit()
            conn.close()
            return True
        except sqlite3.Error as e:
            print(f"Error saving data: {e}")
            return False

# TIMETABLE SOLVER

class TimetableSolver:
    """
    Advanced constraint-based timetable generator using backtracking
    Handles labs, teacher conflicts, room conflicts, and preferences
    """
    
    def __init__(self, classes_data, days, periods):
        self.classes = classes_data
        self.days = days
        self.periods = periods
        
        # Global conflict tracking
        self.teacher_schedule = {d: [set() for _ in range(periods)] for d in range(1, days + 1)}
        self.room_schedule = {d: [set() for _ in range(periods)] for d in range(1, days + 1)}
        
        # Track remaining periods for each subject
        self.remaining_periods = {}
        for class_name, class_info in classes_data.items():
            self.remaining_periods[class_name] = {}
            for subject_name, subject_data in class_info['subjects'].items():
                credit = subject_data.get('credit', 0)
                periods_needed = credit_to_periods(credit)
                self.remaining_periods[class_name][subject_name] = periods_needed
        
        # Initialize empty timetable
        self.timetable = {
            class_name: {
                f"Day {d}": [("FREE HOUR", "-", "-") for _ in range(periods)]
                for d in range(1, days + 1)
            }
            for class_name in classes_data
        }
    
    def get_prioritized_subjects(self, class_name):
        """Get subjects ordered by constraint priority (most constrained first)"""
        subjects = []
        for subject, remaining in self.remaining_periods[class_name].items():
            if remaining > 0:
                subject_data = self.classes[class_name]['subjects'][subject]
                # Priority: labs first, then by remaining periods
                priority = (
                    -int(subject_data.get('lab', False)),
                    -remaining
                )
                subjects.append((priority, subject))
        
        subjects.sort()
        return [s[1] for s in subjects]
    
    def can_place_subject(self, class_name, day, period, subject_name):
        """Check if subject can be placed at given slot"""
        if subject_name == "FREE HOUR":
            return True
        
        subject_data = self.classes[class_name]['subjects'][subject_name]
        teacher = subject_data.get('teacher')
        rooms = subject_data.get('rooms', [])
        
        # Check teacher availability
        if teacher and teacher in self.teacher_schedule[day][period]:
            return False
        
        # Check room availability
        if rooms:
            available_room = any(
                room not in self.room_schedule[day][period]
                for room in rooms
            )
            if not available_room:
                return False
        
        # Check if subject already scheduled today
        day_schedule = self.timetable[class_name][f"Day {day}"]
        if any(slot[0] == subject_name for slot in day_schedule):
            return False
        
        # Check morning/afternoon preferences
        mid_period = self.periods // 2
        if subject_name in self.classes[class_name].get('morning', set()):
            if period >= mid_period:
                return False
        if subject_name in self.classes[class_name].get('afternoon', set()):
            if period < mid_period:
                return False
        
        return True
    
    def place_subject(self, class_name, day, period, subject_name):
        """Place subject at given slot"""
        if subject_name == "FREE HOUR":
            self.timetable[class_name][f"Day {day}"][period] = ("FREE HOUR", "-", "-")
            return True
        
        subject_data = self.classes[class_name]['subjects'][subject_name]
        teacher = subject_data.get('teacher', '')
        rooms = subject_data.get('rooms', [])
        
        # Select available room
        selected_room = "-"
        for room in rooms:
            if room not in self.room_schedule[day][period]:
                selected_room = room
                break
        
        if not selected_room and rooms:
            return False
        
        # Place the subject
        self.timetable[class_name][f"Day {day}"][period] = (
            subject_name, teacher, selected_room
        )
        
        # Update conflict trackers
        if teacher:
            self.teacher_schedule[day][period].add(teacher)
        if selected_room != "-":
            self.room_schedule[day][period].add(selected_room)
        
        # Decrement remaining periods
        self.remaining_periods[class_name][subject_name] -= 1
        
        return True
    
    def remove_subject(self, class_name, day, period, subject_name):
        """Remove subject from given slot"""
        if subject_name == "FREE HOUR":
            return
        
        slot = self.timetable[class_name][f"Day {day}"][period]
        teacher = slot[1]
        room = slot[2]
        
        # Clear conflict trackers
        if teacher != "-":
            self.teacher_schedule[day][period].discard(teacher)
        if room != "-":
            self.room_schedule[day][period].discard(room)
        
        # Restore remaining periods
        self.remaining_periods[class_name][subject_name] += 1
        
        # Clear slot
        self.timetable[class_name][f"Day {day}"][period] = ("FREE HOUR", "-", "-")
    
    def place_lab_continuous(self, class_name, day, start_period, subject_name, length):
        """Place lab subject across continuous periods"""
        # Check if all periods are available
        if start_period + length > self.periods:
            return False
        
        for offset in range(length):
            if not self.can_place_subject(class_name, day, start_period + offset, subject_name):
                return False
        
        # Place across all periods
        for offset in range(length):
            self.place_subject(class_name, day, start_period + offset, subject_name)
        
        return True
    
    def remove_lab_continuous(self, class_name, day, start_period, subject_name, length):
        """Remove lab subject from continuous periods"""
        for offset in range(length):
            self.remove_subject(class_name, day, start_period + offset, subject_name)
    
    def solve_for_class(self, class_name):
        """Solve timetable for a single class using backtracking"""
        def backtrack(day, period):
            # Base case: all slots filled
            if day > self.days:
                return True
            
            # Calculate next slot
            next_day, next_period = (day, period + 1) if period + 1 < self.periods else (day + 1, 0)
            
            # Get prioritized subjects
            candidates = self.get_prioritized_subjects(class_name)
            candidates.append("FREE HOUR")
            
            for subject in candidates:
                if subject != "FREE HOUR" and self.remaining_periods[class_name].get(subject, 0) <= 0:
                    continue
                
                subject_data = self.classes[class_name]['subjects'].get(subject, {})
                
                # Handle lab subjects
                if subject != "FREE HOUR" and subject_data.get('lab'):
                    lab_length = subject_data.get('lab_continuous', 1)
                    if self.place_lab_continuous(class_name, day, period, subject, lab_length):
                        if backtrack(next_day, next_period):
                            return True
                        self.remove_lab_continuous(class_name, day, period, subject, lab_length)
                    continue
                
                # Handle regular subjects
                if not self.can_place_subject(class_name, day, period, subject):
                    continue
                
                self.place_subject(class_name, day, period, subject)
                
                if backtrack(next_day, next_period):
                    return True
                
                self.remove_subject(class_name, day, period, subject)
            
            return False
        
        return backtrack(1, 0)
    
    def solve(self, progress_callback=None):
        """Solve timetables for all classes"""
        # Sort classes by total periods needed (most constrained first)
        sorted_classes = sorted(
            self.classes.keys(),
            key=lambda c: -sum(
                credit_to_periods(s.get('credit', 0))
                for s in self.classes[c]['subjects'].values()
            )
        )
        
        for idx, class_name in enumerate(sorted_classes):
            success = self.solve_for_class(class_name)
            
            if progress_callback:
                progress_callback(idx + 1, len(sorted_classes), class_name, success)
            
            if not success:
                return False
        
        return True

# EXPORT UTILITIES

def export_excel(timetables, teacher_schedules, config):
    """Export timetables to Excel file"""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl library not installed")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    periods = config['periods']
    period_times = config['period_times']
    
    # Create sheets for each class
    for class_name, timetable in timetables.items():
        ws = wb.create_sheet(class_name[:31])
        
        # Header row
        header = ["Day"] + [
            f"P{i+1}\n{period_times[i] if i < len(period_times) else ''}"
            for i in range(periods)
        ]
        ws.append(header)
        
        # Data rows
        for day_key, slots in sorted(timetable.items()):
            row = [day_key] + [
                f"{slot[0]}\n{slot[1]}\n{slot[2]}"
                if slot[0] != 'FREE HOUR' else 'FREE'
                for slot in slots
            ]
            ws.append(row)
        
        # Apply formatting
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Apply colors
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(2, periods + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and 'FREE' not in cell.value:
                    subject_name = cell.value.split('\n')[0]
                    color = config['colors'].get(subject_name, '#FFFFFF')
                    rgb = hex_to_rgb(color)
                    cell.fill = PatternFill(
                        start_color=color.lstrip('#'),
                        end_color=color.lstrip('#'),
                        fill_type='solid'
                    )
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, periods + 2):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Teacher schedule sheet
    if teacher_schedules:
        ws = wb.create_sheet("Teacher Schedules")
        row_idx = 1
        
        for teacher, schedule in sorted(teacher_schedules.items()):
            ws.cell(row=row_idx, column=1, value=f"Teacher: {teacher}").font = Font(bold=True, size=12)
            row_idx += 1
            
            # Header
            header = ["Day"] + [f"P{i+1}" for i in range(periods)]
            for col_idx, val in enumerate(header, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            row_idx += 1
            
            # Data
            for day_key, slots in sorted(schedule.items()):
                ws.cell(row=row_idx, column=1, value=day_key)
                for col_idx, slot in enumerate(slots, 2):
                    val = f"{slot[0]}\n{slot[1]},{slot[2]}" if slot[0] != 'FREE HOUR' else 'FREE'
                    ws.cell(row=row_idx, column=col_idx, value=val)
                row_idx += 1
            
            row_idx += 1
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = os.path.join(EXPORTS_DIR, f"timetables_{timestamp}.xlsx")
    wb.save(filename)
    
    return filename

def export_pdf(timetables, config):
    """Export timetables to PDF file"""
    if not HAS_REPORTLAB:
        raise RuntimeError("reportlab library not installed")
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = os.path.join(EXPORTS_DIR, f"timetables_{timestamp}.pdf")
    
    c = pdf_canvas.Canvas(filename, pagesize=landscape(A4))
    width, height = landscape(A4)
    
    margin = 40
    y_pos = height - margin
    
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y_pos, "College Timetables")
    y_pos -= 30
    
    for class_name, timetable in timetables.items():
        if y_pos < 100:
            c.showPage()
            y_pos = height - margin
        
        c.setFont('Helvetica-Bold', 14)
        c.drawString(margin, y_pos, f"Class: {class_name}")
        y_pos -= 20
        
        c.setFont('Helvetica', 10)
        for day_key, slots in sorted(timetable.items()):
            line = f"{day_key}: " + " | ".join([slot[0] for slot in slots])
            c.drawString(margin, y_pos, line[:100])
            y_pos -= 15
            
            if y_pos < 60:
                c.showPage()
                y_pos = height - margin
        
        y_pos -= 10
    
    c.save()
    return filename

def aggregate_teacher_schedules(timetables, periods, days):
    """Aggregate teacher schedules from all class timetables"""
    teacher_map = defaultdict(lambda: {
        f"Day {d}": [('FREE HOUR', '-', '-') for _ in range(periods)]
        for d in range(1, days + 1)
    })
    
    for class_name, timetable in timetables.items():
        for day_key, slots in timetable.items():
            for period_idx, (subject, teacher, room) in enumerate(slots):
                if subject != 'FREE HOUR' and teacher != '-':
                    teacher_map[teacher][day_key][period_idx] = (subject, room, class_name)
    
    return dict(teacher_map)

# CUSTOM DIALOGS

class SubjectDialog(tk.Toplevel):
    """Custom dialog for adding/editing subjects"""
    
    def __init__(self, parent, title="Subject Details", initial_data=None):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.initial_data = initial_data or {}
        
        self.create_widgets()
        self.center_window()
        self.transient(parent)
        self.grab_set()
    
    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Subject Name
        ttk.Label(main_frame, text="Subject Name:").grid(row=0, column=0, sticky='w', pady=5)
        self.name_var = tk.StringVar(value=self.initial_data.get('name', ''))
        ttk.Entry(main_frame, textvariable=self.name_var, width=30).grid(row=0, column=1, pady=5)
        
        # Teacher
        ttk.Label(main_frame, text="Teacher:").grid(row=1, column=0, sticky='w', pady=5)
        self.teacher_var = tk.StringVar(value=self.initial_data.get('teacher', ''))
        ttk.Entry(main_frame, textvariable=self.teacher_var, width=30).grid(row=1, column=1, pady=5)
        
        # Credit
        ttk.Label(main_frame, text="Credit Hours:").grid(row=2, column=0, sticky='w', pady=5)
        self.credit_var = tk.IntVar(value=self.initial_data.get('credit', 3))
        ttk.Spinbox(main_frame, from_=1, to=10, textvariable=self.credit_var, width=28).grid(row=2, column=1, pady=5)
        
        # Lab checkbox
        self.lab_var = tk.BooleanVar(value=self.initial_data.get('lab', False))
        ttk.Checkbutton(main_frame, text="Is Lab Subject", variable=self.lab_var, command=self.toggle_lab_options).grid(row=3, column=0, columnspan=2, sticky='w', pady=5)
        
        # Lab continuous periods
        ttk.Label(main_frame, text="Continuous Periods:").grid(row=4, column=0, sticky='w', pady=5)
        self.lab_cont_var = tk.IntVar(value=self.initial_data.get('lab_continuous', 2))
        self.lab_cont_spin = ttk.Spinbox(main_frame, from_=1, to=4, textvariable=self.lab_cont_var, width=28)
        self.lab_cont_spin.grid(row=4, column=1, pady=5)
        
        # Rooms
        ttk.Label(main_frame, text="Rooms (comma-separated):").grid(row=5, column=0, sticky='w', pady=5)
        rooms_str = ','.join(self.initial_data.get('rooms', []))
        self.rooms_var = tk.StringVar(value=rooms_str)
        ttk.Entry(main_frame, textvariable=self.rooms_var, width=30).grid(row=5, column=1, pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="OK", command=self.ok_clicked).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.cancel_clicked).pack(side='left', padx=5)
        
        self.toggle_lab_options()
    
    def toggle_lab_options(self):
        """Enable/disable lab options based on checkbox"""
        state = 'normal' if self.lab_var.get() else 'disabled'
        self.lab_cont_spin.config(state=state)
    
    def center_window(self):
        """Center the dialog on screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'+{x}+{y}')
    
    def ok_clicked(self):
        """Validate and save data"""
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror("Error", "Subject name is required", parent=self)
            return
        
        rooms_str = self.rooms_var.get().strip()
        rooms = [r.strip() for r in rooms_str.split(',') if r.strip()]
        
        self.result = {
            'name': name,
            'teacher': self.teacher_var.get().strip(),
            'credit': self.credit_var.get(),
            'lab': self.lab_var.get(),
            'lab_continuous': self.lab_cont_var.get() if self.lab_var.get() else 1,
            'rooms': rooms
        }
        
        self.destroy()
    
    def cancel_clicked(self):
        """Cancel dialog"""
        self.result = None
        self.destroy()

# MAIN APPLICATION

class TimetableApp(tk.Tk):
    """Main application window with modern UI"""
    
    def __init__(self):
        super().__init__()
        
        self.title("Advanced Timetable Generator v2.0")
        self.geometry("1400x900")
        self.configure(bg=UI_COLORS['bg_main'])
        
        # Initialize managers
        self.db_manager = DatabaseManager()
        
        # Application state
        self.state = {
            'classes': {},
            'colors': {},
            'timetables': {},
            'teacher_schedules': {},
            'days': 5,
            'periods': 8,
            'period_times': ["8:00-8:50", "9:00-9:50", "10:00-10:50", "11:00-11:50",
                           "12:00-12:50", "1:00-1:50", "2:00-2:50", "3:00-3:50"]
        }
        
        self.solver_thread = None
        self.selected_class = None
        
        # Setup UI
        self.setup_styles()
        self.create_widgets()
        self.load_data()
        
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def setup_styles(self):
        """Configure ttk styles for modern look"""
        style = ttk.Style()
        
        # Try to use a modern theme
        available_themes = style.theme_names()
        if 'clam' in available_themes:
            style.theme_use('clam')
        
        # Configure custom styles
        style.configure('Title.TLabel',
                       font=('Helvetica', 16, 'bold'),
                       foreground=UI_COLORS['primary'],
                       background=UI_COLORS['bg_main'])
        
        style.configure('Subtitle.TLabel',
                       font=('Helvetica', 11, 'bold'),
                       foreground=UI_COLORS['text_primary'],
                       background=UI_COLORS['bg_secondary'])
        
        style.configure('Action.TButton',
                       font=('Helvetica', 10, 'bold'),
                       padding=10)
        
        style.configure('Success.TButton',
                       foreground=UI_COLORS['success'])
        
        style.configure('Danger.TButton',
                       foreground=UI_COLORS['danger'])
    
    def create_widgets(self):
        """Create all UI widgets"""
        # Main container with padding
        main_container = ttk.Frame(self, padding="10")
        main_container.pack(fill='both', expand=True)
        
        # Title bar
        self.create_title_bar(main_container)
        
        # Toolbar
        self.create_toolbar(main_container)
        
        # Main content area with paned window
        content_pane = ttk.PanedWindow(main_container, orient='horizontal')
        content_pane.pack(fill='both', expand=True, pady=(10, 0))
        
        # Left panel - Class and Subject Management
        left_panel = self.create_left_panel(content_pane)
        content_pane.add(left_panel, weight=1)
        
        # Right panel - Timetable Preview
        right_panel = self.create_right_panel(content_pane)
        content_pane.add(right_panel, weight=2)
        
        # Status bar
        self.create_status_bar(main_container)
    
    def create_title_bar(self, parent):
        """Create application title bar"""
        title_frame = tk.Frame(parent, bg=UI_COLORS['primary'], height=60)
        title_frame.pack(fill='x', pady=(0, 10))
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(
            title_frame,
            text="🎓 KIT Timetable Generator",
            font=('Helvetica', 20, 'bold'),
            bg=UI_COLORS['primary'],
            fg='white'
        )
        title_label.pack(pady=15)
    
    def create_toolbar(self, parent):
        """Create toolbar with main actions"""
        toolbar = tk.Frame(parent, bg=UI_COLORS['bg_secondary'], relief='raised', bd=1)
        toolbar.pack(fill='x', pady=(0, 10))
        
        # File operations
        file_frame = ttk.LabelFrame(toolbar, text="File Operations", padding="5")
        file_frame.pack(side='left', padx=5, pady=5)
        
        ttk.Button(file_frame, text="💾 Save", command=self.save_data, width=10).pack(side='left', padx=2)
        ttk.Button(file_frame, text="📂 Load", command=self.load_data, width=10).pack(side='left', padx=2)
        
        # Export operations
        export_frame = ttk.LabelFrame(toolbar, text="Export", padding="5")
        export_frame.pack(side='left', padx=5, pady=5)
        
        ttk.Button(export_frame, text="📊 Excel", command=self.export_excel, width=10).pack(side='left', padx=2)
        ttk.Button(export_frame, text="📄 PDF", command=self.export_pdf, width=10).pack(side='left', padx=2)
        ttk.Button(export_frame, text="💾 JSON", command=self.export_json, width=10).pack(side='left', padx=2)
        
        # Configuration
        config_frame = ttk.LabelFrame(toolbar, text="Configuration", padding="5")
        config_frame.pack(side='left', padx=5, pady=5)
        
        ttk.Label(config_frame, text="Days:").pack(side='left', padx=2)
        self.days_var = tk.IntVar(value=self.state['days'])
        ttk.Spinbox(config_frame, from_=5, to=7, textvariable=self.days_var, width=5).pack(side='left', padx=2)
        
        ttk.Label(config_frame, text="Periods:").pack(side='left', padx=5)
        self.periods_var = tk.IntVar(value=self.state['periods'])
        ttk.Spinbox(config_frame, from_=6, to=10, textvariable=self.periods_var, width=5).pack(side='left', padx=2)
    
    def create_left_panel(self, parent):
        """Create left panel for class and subject management"""
        panel = ttk.Frame(parent)
        
        # Class Management Section
        class_frame = ttk.LabelFrame(panel, text="📚 Class Management", padding="10")
        class_frame.pack(fill='both', expand=True, pady=(0, 5))
        
        # Add class controls
        add_class_frame = ttk.Frame(class_frame)
        add_class_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(add_class_frame, text="Class Name:").pack(side='left', padx=(0, 5))
        self.class_name_var = tk.StringVar()
        ttk.Entry(add_class_frame, textvariable=self.class_name_var, width=20).pack(side='left', padx=(0, 5))
        ttk.Button(add_class_frame, text="➕ Add Class", command=self.add_class).pack(side='left')
        
        # Class list
        list_frame = ttk.Frame(class_frame)
        list_frame.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.class_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=('Helvetica', 10),
            bg='white',
            selectmode='single',
            height=8
        )
        self.class_listbox.pack(side='left', fill='both', expand=True)
        self.class_listbox.bind('<<ListboxSelect>>', self.on_class_selected)
        scrollbar.config(command=self.class_listbox.yview)
        
        # Class action buttons
        class_btn_frame = ttk.Frame(class_frame)
        class_btn_frame.pack(fill='x', pady=(5, 0))
        
        ttk.Button(class_btn_frame, text="🗑️ Delete Class", command=self.delete_class).pack(side='left', padx=2)
        
        # Subject Management Section
        subject_frame = ttk.LabelFrame(panel, text="📖 Subject Management", padding="10")
        subject_frame.pack(fill='both', expand=True)
        
        # Subject list with search
        search_frame = ttk.Frame(subject_frame)
        search_frame.pack(fill='x', pady=(0, 5))
        
        ttk.Label(search_frame, text="🔍 Search:").pack(side='left', padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_subjects())
        ttk.Entry(search_frame, textvariable=self.search_var, width=25).pack(side='left', fill='x', expand=True)
        
        # Subject listbox
        subj_list_frame = ttk.Frame(subject_frame)
        subj_list_frame.pack(fill='both', expand=True, pady=(0, 5))
        
        subj_scrollbar = ttk.Scrollbar(subj_list_frame)
        subj_scrollbar.pack(side='right', fill='y')
        
        self.subject_listbox = tk.Listbox(
            subj_list_frame,
            yscrollcommand=subj_scrollbar.set,
            font=('Helvetica', 9),
            bg='white',
            selectmode='single',
            height=12
        )
        self.subject_listbox.pack(side='left', fill='both', expand=True)
        subj_scrollbar.config(command=self.subject_listbox.yview)
        
        # Subject action buttons
        subj_btn_frame = ttk.Frame(subject_frame)
        subj_btn_frame.pack(fill='x')
        
        ttk.Button(subj_btn_frame, text="➕ Add", command=self.add_subject, width=10).pack(side='left', padx=2)
        ttk.Button(subj_btn_frame, text="✏️ Edit", command=self.edit_subject, width=10).pack(side='left', padx=2)
        ttk.Button(subj_btn_frame, text="🗑️ Delete", command=self.delete_subject, width=10).pack(side='left', padx=2)
        ttk.Button(subj_btn_frame, text="🎨 Color", command=self.assign_color, width=10).pack(side='left', padx=2)
        
        return panel
    
    def create_right_panel(self, parent):
        """Create right panel for timetable generation and preview"""
        panel = ttk.Frame(parent)
        
        # Generation controls
        gen_frame = ttk.LabelFrame(panel, text="⚙️ Timetable Generation", padding="10")
        gen_frame.pack(fill='x', pady=(0, 10))
        
        btn_frame = ttk.Frame(gen_frame)
        btn_frame.pack(fill='x')
        
        ttk.Button(
            btn_frame,
            text="🚀 Generate Timetables",
            command=self.generate_timetables,
            style='Action.TButton'
        ).pack(side='left', padx=5)
        
        ttk.Button(
            btn_frame,
            text="🗑️ Clear All",
            command=self.clear_timetables
        ).pack(side='left', padx=5)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            gen_frame,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.pack(fill='x', pady=(10, 5))
        
        # Log text
        log_frame = ttk.LabelFrame(gen_frame, text="Generation Log", padding="5")
        log_frame.pack(fill='x', pady=(5, 0))
        
        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side='right', fill='y')
        
        self.log_text = tk.Text(
            log_frame,
            height=5,
            font=('Courier', 9),
            bg='#F5F5F5',
            yscrollcommand=log_scroll.set,
            wrap='word'
        )
        self.log_text.pack(fill='x')
        log_scroll.config(command=self.log_text.yview)
        
        # Preview area with tabs
        preview_frame = ttk.LabelFrame(panel, text="📅 Timetable Preview", padding="10")
        preview_frame.pack(fill='both', expand=True)
        
        self.preview_notebook = ttk.Notebook(preview_frame)
        self.preview_notebook.pack(fill='both', expand=True)
        
        # Class timetables tab
        self.class_preview_frame = ttk.Frame(self.preview_notebook)
        self.preview_notebook.add(self.class_preview_frame, text="Class Timetables")
        
        # Teacher schedules tab
        self.teacher_preview_frame = ttk.Frame(self.preview_notebook)
        self.preview_notebook.add(self.teacher_preview_frame, text="Teacher Schedules")
        
        return panel
    
    def create_status_bar(self, parent):
        """Create status bar at bottom"""
        self.status_frame = tk.Frame(parent, bg=UI_COLORS['secondary'], height=25, relief='sunken', bd=1)
        self.status_frame.pack(fill='x', side='bottom')
        self.status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(
            self.status_frame,
            text="Ready",
            bg=UI_COLORS['secondary'],
            fg='white',
            font=('Helvetica', 9),
            anchor='w',
            padx=10
        )
        self.status_label.pack(fill='x')
    
    def update_status(self, message):
        """Update status bar message"""
        self.status_label.config(text=message)
        self.log_message(message)
    
    def log_message(self, message):
        """Add message to log"""
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        self.log_text.insert('end', f"[{timestamp}] {message}\n")
        self.log_text.see('end')
    
    # Class Management Methods
    
    def add_class(self):
        """Add a new class"""
        class_name = self.class_name_var.get().strip()
        
        if not class_name:
            messagebox.showwarning("Input Required", "Please enter a class name")
            return
        
        if class_name in self.state['classes']:
            messagebox.showinfo("Duplicate", f"Class '{class_name}' already exists")
            return
        
        self.state['classes'][class_name] = {
            'subjects': {},
            'morning': set(),
            'afternoon': set()
        }
        
        self.class_listbox.insert('end', class_name)
        self.class_name_var.set('')
        self.update_status(f"Added class: {class_name}")
    
    def delete_class(self):
        """Delete selected class"""
        selection = self.class_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a class to delete")
            return
        
        class_name = self.class_listbox.get(selection[0])
        
        if messagebox.askyesno("Confirm Delete", f"Delete class '{class_name}' and all its subjects?"):
            del self.state['classes'][class_name]
            self.class_listbox.delete(selection[0])
            self.subject_listbox.delete(0, 'end')
            self.selected_class = None
            self.update_status(f"Deleted class: {class_name}")
    
    def on_class_selected(self, event):
        """Handle class selection"""
        selection = self.class_listbox.curselection()
        if selection:
            self.selected_class = self.class_listbox.get(selection[0])
            self.refresh_subject_list()
    
    # Subject Management Methods
    
    def add_subject(self):
        """Add a new subject to selected class"""
        if not self.selected_class:
            messagebox.showwarning("No Class", "Please select a class first")
            return
        
        dialog = SubjectDialog(self, "Add Subject")
        self.wait_window(dialog)
        
        if dialog.result:
            subject_name = dialog.result['name']
            
            if subject_name in self.state['classes'][self.selected_class]['subjects']:
                messagebox.showinfo("Duplicate", f"Subject '{subject_name}' already exists")
                return
            
            self.state['classes'][self.selected_class]['subjects'][subject_name] = {
                'teacher': dialog.result['teacher'],
                'credit': dialog.result['credit'],
                'lab': dialog.result['lab'],
                'lab_continuous': dialog.result['lab_continuous'],
                'rooms': dialog.result['rooms']
            }
            
            # Assign color if not exists
            if subject_name not in self.state['colors']:
                self.state['colors'][subject_name] = self.get_next_color()
            
            self.refresh_subject_list()
            self.update_status(f"Added subject: {subject_name}")
    
    def edit_subject(self):
        """Edit selected subject"""
        selection = self.subject_listbox.curselection()
        if not selection or not self.selected_class:
            messagebox.showwarning("No Selection", "Please select a subject to edit")
            return
        
        subject_line = self.subject_listbox.get(selection[0])
        subject_name = subject_line.split('|')[0].strip()
        
        current_data = self.state['classes'][self.selected_class]['subjects'][subject_name].copy()
        current_data['name'] = subject_name
        
        dialog = SubjectDialog(self, "Edit Subject", current_data)
        self.wait_window(dialog)
        
        if dialog.result:
            # Remove old entry if name changed
            if dialog.result['name'] != subject_name:
                del self.state['classes'][self.selected_class]['subjects'][subject_name]
                subject_name = dialog.result['name']
            
            self.state['classes'][self.selected_class]['subjects'][subject_name] = {
                'teacher': dialog.result['teacher'],
                'credit': dialog.result['credit'],
                'lab': dialog.result['lab'],
                'lab_continuous': dialog.result['lab_continuous'],
                'rooms': dialog.result['rooms']
            }
            
            self.refresh_subject_list()
            self.update_status(f"Updated subject: {subject_name}")
    
    def delete_subject(self):
        """Delete selected subject"""
        selection = self.subject_listbox.curselection()
        if not selection or not self.selected_class:
            messagebox.showwarning("No Selection", "Please select a subject to delete")
            return
        
        subject_line = self.subject_listbox.get(selection[0])
        subject_name = subject_line.split('|')[0].strip()
        
        if messagebox.askyesno("Confirm Delete", f"Delete subject '{subject_name}'?"):
            del self.state['classes'][self.selected_class]['subjects'][subject_name]
            self.refresh_subject_list()
            self.update_status(f"Deleted subject: {subject_name}")
    
    def assign_color(self):
        """Assign color to selected subject"""
        selection = self.subject_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a subject")
            return
        
        subject_line = self.subject_listbox.get(selection[0])
        subject_name = subject_line.split('|')[0].strip()
        
        current_color = self.state['colors'].get(subject_name, '#FFFFFF')
        color = colorchooser.askcolor(title=f"Choose color for {subject_name}", initialcolor=current_color)
        
        if color and color[1]:
            self.state['colors'][subject_name] = color[1]
            self.refresh_subject_list()
            self.update_status(f"Color assigned to: {subject_name}")
    
    def refresh_subject_list(self):
        """Refresh the subject listbox"""
        self.subject_listbox.delete(0, 'end')
        
        if not self.selected_class:
            return
        
        search_term = self.search_var.get().lower()
        subjects = self.state['classes'][self.selected_class]['subjects']
        
        for name, data in sorted(subjects.items()):
            display_text = f"{name} | {data['teacher']} | {data['credit']} credits"
            if data['lab']:
                display_text += " | LAB"
            
            if search_term and search_term not in display_text.lower():
                continue
            
            self.subject_listbox.insert('end', display_text)
            
            # Color the entry
            if name in self.state['colors']:
                idx = self.subject_listbox.size() - 1
                self.subject_listbox.itemconfig(idx, bg=self.state['colors'][name])
    
    def filter_subjects(self):
        """Filter subjects based on search"""
        self.refresh_subject_list()
    
    def get_next_color(self):
        """Get next available color from palette"""
        used_colors = set(self.state['colors'].values())
        available = [c for c in COLOR_PALETTE if c not in used_colors]
        return available[0] if available else random.choice(COLOR_PALETTE)
    
    # Timetable Generation Methods
    
    def generate_timetables(self):
        """Generate timetables for all classes"""
        if not self.state['classes']:
            messagebox.showwarning("No Data", "Please add classes and subjects first")
            return
        
        if self.solver_thread and self.solver_thread.is_alive():
            messagebox.showinfo("Busy", "Generation is already in progress")
            return
        
        # Update configuration
        self.state['days'] = self.days_var.get()
        self.state['periods'] = self.periods_var.get()
        
        self.progress_var.set(0)
        self.log_text.delete('1.0', 'end')
        self.update_status("Starting timetable generation...")
        
        # Run solver in separate thread
        self.solver_thread = threading.Thread(target=self.run_solver, daemon=True)
        self.solver_thread.start()
    
    def run_solver(self):
        """Run the timetable solver"""
        try:
            solver = TimetableSolver(
                deepcopy(self.state['classes']),
                self.state['days'],
                self.state['periods']
            )
            
            def progress_callback(current, total, class_name, success):
                progress = (current / total) * 100
                self.progress_var.set(progress)
                status = "✓ Success" if success else "✗ Failed"
                self.log_message(f"{class_name}: {status} ({current}/{total})")
            
            success = solver.solve(progress_callback)
            
            if success:
                self.state['timetables'] = solver.timetable
                self.state['teacher_schedules'] = aggregate_teacher_schedules(
                    solver.timetable,
                    self.state['periods'],
                    self.state['days']
                )
                self.update_status("✓ Timetables generated successfully!")
                self.after(0, self.refresh_preview)
                messagebox.showinfo("Success", "Timetables generated successfully!")
            else:
                self.update_status("✗ Generation failed - please adjust constraints")
                messagebox.showwarning("Failed", "Could not generate complete timetables. Try adjusting subject credits or periods.")
        
        except Exception as e:
            self.update_status(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Generation error: {str(e)}")
    
    def clear_timetables(self):
        """Clear generated timetables"""
        if messagebox.askyesno("Confirm", "Clear all generated timetables?"):
            self.state['timetables'] = {}
            self.state['teacher_schedules'] = {}
            self.refresh_preview()
            self.update_status("Timetables cleared")
    
    def refresh_preview(self):
        """Refresh timetable preview"""
        # Clear existing preview
        for widget in self.class_preview_frame.winfo_children():
            widget.destroy()
        for widget in self.teacher_preview_frame.winfo_children():
            widget.destroy()
        
        if not self.state['timetables']:
            ttk.Label(self.class_preview_frame, text="No timetables generated yet", font=('Helvetica', 12)).pack(pady=50)
            return
        
        # Class timetables preview
        self.render_class_preview()
        
        # Teacher schedules preview
        self.render_teacher_preview()
    
    def render_class_preview(self):
        """Render class timetables preview"""
        if not self.state['timetables']:
            return
        
        # Class selector
        selector_frame = ttk.Frame(self.class_preview_frame)
        selector_frame.pack(fill='x', pady=5)
        
        ttk.Label(selector_frame, text="Select Class:").pack(side='left', padx=5)
        
        class_names = list(self.state['timetables'].keys())
        self.preview_class_var = tk.StringVar(value=class_names[0] if class_names else "")
        
        class_combo = ttk.Combobox(
            selector_frame,
            textvariable=self.preview_class_var,
            values=class_names,
            state='readonly',
            width=20
        )
        class_combo.pack(side='left', padx=5)
        class_combo.bind('<<ComboboxSelected>>', lambda e: self.update_class_preview())
        
        # Timetable grid container
        self.class_grid_container = ttk.Frame(self.class_preview_frame)
        self.class_grid_container.pack(fill='both', expand=True, pady=5)
        
        self.update_class_preview()
    
    def update_class_preview(self):
        """Update the class timetable grid"""
        for widget in self.class_grid_container.winfo_children():
            widget.destroy()
        
        class_name = self.preview_class_var.get()
        if not class_name or class_name not in self.state['timetables']:
            return
        
        timetable = self.state['timetables'][class_name]
        
        # Create scrollable canvas
        canvas = tk.Canvas(self.class_grid_container, bg='white')
        scrollbar_y = ttk.Scrollbar(self.class_grid_container, orient='vertical', command=canvas.yview)
        scrollbar_x = ttk.Scrollbar(self.class_grid_container, orient='horizontal', command=canvas.xview)
        
        grid_frame = ttk.Frame(canvas)
        
        canvas.create_window((0, 0), window=grid_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Header row
        tk.Label(grid_frame, text="Day/Period", relief='solid', width=12, height=2, bg='#34495E', fg='white', font=('Helvetica', 9, 'bold')).grid(row=0, column=0, sticky='nsew')
        
        for p in range(self.state['periods']):
            period_label = f"Period {p+1}"
            if p < len(self.state['period_times']):
                period_label += f"\n{self.state['period_times'][p]}"
            tk.Label(grid_frame, text=period_label, relief='solid', width=14, height=2, bg='#34495E', fg='white', font=('Helvetica', 8, 'bold')).grid(row=0, column=p+1, sticky='nsew')
        
        # Data rows
        row = 1
        for day_key, slots in sorted(timetable.items(), key=lambda x: int(x[0].split()[1])):
            tk.Label(grid_frame, text=day_key, relief='solid', width=12, height=4, bg='#ECF0F1', font=('Helvetica', 9, 'bold')).grid(row=row, column=0, sticky='nsew')
            
            for col, (subject, teacher, room) in enumerate(slots):
                if subject == 'FREE HOUR':
                    text = "FREE"
                    bg_color = '#F8F9FA'
                else:
                    text = f"{subject}\n{teacher}\n{room}"
                    bg_color = self.state['colors'].get(subject, '#FFFFFF')
                
                cell = tk.Label(
                    grid_frame,
                    text=text,
                    relief='solid',
                    width=14,
                    height=4,
                    bg=bg_color,
                    font=('Helvetica', 8),
                    wraplength=100,
                    justify='center'
                )
                cell.grid(row=row, column=col+1, sticky='nsew')
            
            row += 1
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        grid_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox('all'))
    
    def render_teacher_preview(self):
        """Render teacher schedules preview"""
        if not self.state['teacher_schedules']:
            return
        
        # Teacher selector
        selector_frame = ttk.Frame(self.teacher_preview_frame)
        selector_frame.pack(fill='x', pady=5)
        
        ttk.Label(selector_frame, text="Select Teacher:").pack(side='left', padx=5)
        
        teacher_names = list(self.state['teacher_schedules'].keys())
        self.preview_teacher_var = tk.StringVar(value=teacher_names[0] if teacher_names else "")
        
        teacher_combo = ttk.Combobox(
            selector_frame,
            textvariable=self.preview_teacher_var,
            values=teacher_names,
            state='readonly',
            width=20
        )
        teacher_combo.pack(side='left', padx=5)
        teacher_combo.bind('<<ComboboxSelected>>', lambda e: self.update_teacher_preview())
        
        # Schedule grid container
        self.teacher_grid_container = ttk.Frame(self.teacher_preview_frame)
        self.teacher_grid_container.pack(fill='both', expand=True, pady=5)
        
        self.update_teacher_preview()
    
    def update_teacher_preview(self):
        """Update the teacher schedule grid"""
        for widget in self.teacher_grid_container.winfo_children():
            widget.destroy()
        
        teacher_name = self.preview_teacher_var.get()
        if not teacher_name or teacher_name not in self.state['teacher_schedules']:
            return
        
        schedule = self.state['teacher_schedules'][teacher_name]
        
        # Create scrollable canvas
        canvas = tk.Canvas(self.teacher_grid_container, bg='white')
        scrollbar_y = ttk.Scrollbar(self.teacher_grid_container, orient='vertical', command=canvas.yview)
        scrollbar_x = ttk.Scrollbar(self.teacher_grid_container, orient='horizontal', command=canvas.xview)
        
        grid_frame = ttk.Frame(canvas)
        
        canvas.create_window((0, 0), window=grid_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Header row
        tk.Label(grid_frame, text="Day/Period", relief='solid', width=12, height=2, bg='#E74C3C', fg='white', font=('Helvetica', 9, 'bold')).grid(row=0, column=0, sticky='nsew')
        
        for p in range(self.state['periods']):
            period_label = f"Period {p+1}"
            tk.Label(grid_frame, text=period_label, relief='solid', width=14, height=2, bg='#E74C3C', fg='white', font=('Helvetica', 8, 'bold')).grid(row=0, column=p+1, sticky='nsew')
        
        # Data rows
        row = 1
        for day_key, slots in sorted(schedule.items(), key=lambda x: int(x[0].split()[1])):
            tk.Label(grid_frame, text=day_key, relief='solid', width=12, height=4, bg='#FADBD8', font=('Helvetica', 9, 'bold')).grid(row=row, column=0, sticky='nsew')
            
            for col, (subject, room, class_name) in enumerate(slots):
                if subject == 'FREE HOUR':
                    text = "FREE"
                    bg_color = '#F8F9FA'
                else:
                    text = f"{subject}\n{class_name}\n{room}"
                    bg_color = self.state['colors'].get(subject, '#FFFFFF')
                
                cell = tk.Label(
                    grid_frame,
                    text=text,
                    relief='solid',
                    width=14,
                    height=4,
                    bg=bg_color,
                    font=('Helvetica', 8),
                    wraplength=100,
                    justify='center'
                )
                cell.grid(row=row, column=col+1, sticky='nsew')
            
            row += 1
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        grid_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox('all'))
    
    # Data Management Methods
    
    
    def load_data(self):
        """Load data from database"""
        try:
            data = self.db_manager.load_all_data()
            self.state['classes'] = data['classes']
            self.state['colors'] = data['colors']
            
            # Refresh UI
            self.class_listbox.delete(0, 'end')
            for class_name in sorted(self.state['classes'].keys()):
                self.class_listbox.insert('end', class_name)
            
            self.update_status(f"Loaded {len(self.state['classes'])} classes from database")
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load data: {str(e)}")
    
    def save_data(self):
        """Save data to database"""
        try:
            data = {
                'classes': self.state['classes'],
                'colors': self.state['colors']
            }
            
            if self.db_manager.save_all_data(data):
                self.update_status("Data saved successfully")
                messagebox.showinfo("Success", "Data saved to database")
            else:
                messagebox.showerror("Error", "Failed to save data")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save data: {str(e)}")
    
    # Export Methods
    
    
    def export_excel(self):
        """Export timetables to Excel"""
        if not self.state['timetables']:
            messagebox.showwarning("No Data", "Generate timetables first")
            return
        
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library", "openpyxl is not installed.\nInstall with: pip install openpyxl")
            return
        
        try:
            config = {
                'periods': self.state['periods'],
                'period_times': self.state['period_times'],
                'colors': self.state['colors']
            }
            
            filename = export_excel(
                self.state['timetables'],
                self.state['teacher_schedules'],
                config
            )
            
            self.update_status(f"Exported to: {filename}")
            messagebox.showinfo("Success", f"Exported to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel: {str(e)}")
    
    def export_pdf(self):
        """Export timetables to PDF"""
        if not self.state['timetables']:
            messagebox.showwarning("No Data", "Generate timetables first")
            return
        
        if not HAS_REPORTLAB:
            messagebox.showerror("Missing Library", "reportlab is not installed.\nInstall with: pip install reportlab")
            return
        
        try:
            config = {
                'periods': self.state['periods'],
                'period_times': self.state['period_times'],
                'colors': self.state['colors']
            }
            
            filename = export_pdf(self.state['timetables'], config)
            
            self.update_status(f"Exported to: {filename}")
            messagebox.showinfo("Success", f"Exported to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export PDF: {str(e)}")
    
    def export_json(self):
        """Export data to JSON file"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension='.json',
                filetypes=[('JSON Files', '*.json'), ('All Files', '*.*')],
                initialdir=EXPORTS_DIR
            )
            
            if not filename:
                return
            
            export_data = {
                'classes': {},
                'colors': self.state['colors'],
                'config': {
                    'days': self.state['days'],
                    'periods': self.state['periods'],
                    'period_times': self.state['period_times']
                }
            }
            
            # Convert sets to lists for JSON serialization
            for class_name, class_data in self.state['classes'].items():
                export_data['classes'][class_name] = {
                    'subjects': class_data['subjects'],
                    'morning': list(class_data.get('morning', [])),
                    'afternoon': list(class_data.get('afternoon', []))
                }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            self.update_status(f"Exported to: {filename}")
            messagebox.showinfo("Success", f"Exported to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export JSON: {str(e)}")
    
    def import_json(self):
        """Import data from JSON file"""
        try:
            filename = filedialog.askopenfilename(
                filetypes=[('JSON Files', '*.json'), ('All Files', '*.*')],
                initialdir=EXPORTS_DIR
            )
            
            if not filename:
                return
            
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Load classes
            self.state['classes'] = {}
            for class_name, class_data in data.get('classes', {}).items():
                self.state['classes'][class_name] = {
                    'subjects': class_data.get('subjects', {}),
                    'morning': set(class_data.get('morning', [])),
                    'afternoon': set(class_data.get('afternoon', []))
                }
            
            # Load colors
            self.state['colors'] = data.get('colors', {})
            
            # Load config if available
            if 'config' in data:
                self.state['days'] = data['config'].get('days', 5)
                self.state['periods'] = data['config'].get('periods', 8)
                self.state['period_times'] = data['config'].get('period_times', self.state['period_times'])
                
                self.days_var.set(self.state['days'])
                self.periods_var.set(self.state['periods'])
            
            # Refresh UI
            self.class_listbox.delete(0, 'end')
            for class_name in sorted(self.state['classes'].keys()):
                self.class_listbox.insert('end', class_name)
            
            self.update_status(f"Imported from: {filename}")
            messagebox.showinfo("Success", "Data imported successfully")
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import JSON: {str(e)}")
    
    # Application Control
  
    
    def on_closing(self):
        """Handle application closing"""
        if messagebox.askokcancel("Quit", "Do you want to save before quitting?"):
            self.save_data()
        self.destroy()




def main():
    
    try:
        app = TimetableApp()
        app.mainloop()
    except Exception as e:
        print(f"Application error: {e}")
        messagebox.showerror("Fatal Error", f"Application failed to start:\n{str(e)}")

if __name__ == '__main__':
    main()